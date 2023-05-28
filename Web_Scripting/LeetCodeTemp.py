from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# URL of the site to fetch data
siteUrl = 'https://leetcode.com/problemset/all/'
questionNameList = []
questionUrlList = []
questionDifficultyList = []


def xcelSheet():

    excelFileName = 'LeetCode.xlsx'
    sheetName = 'LeetCode Problems'

    df = pd.DataFrame({
        'Question Name': questionNameList,
        'Question Url': questionUrlList,
        'Question Difficulty': questionDifficultyList
    })

    wb = Workbook()
    sheet1 = wb.create_sheet(sheetName)
    sheet1.cell(1, 1, 'Question Name')
    sheet1.cell(1, 2, 'Question URL')
    sheet1.cell(1, 3, 'Question Difficulty')

    for i in range(0, df.__len__()):
        sheet1.cell(i + 2, 1, df['Question Name'][i])
        sheet1.cell(i + 2, 2, df['Question Url'][i])
        sheet1.cell(i + 2, 3, df['Question Difficulty'][i])

    wb.save(excelFileName)
    wb.close()
    print(" ****Excel sheet created***** ")


def openBrowser(url):
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument('--incognito')
    options.add_argument('--headless')

    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    # headless browser
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                              options=options)

    driver.get(url)
    driver.maximize_window()
    return driver


def closeBrowser(driver):
    driver.close()


def fetchPageData(pageUrl):
    sleepTime = 3

    # print("Page URL: ", pageUrl)
    browser = openBrowser(pageUrl)
    time.sleep(sleepTime)
    pageSource = browser.page_source
    WebDriverWait(browser, 10).until(EC.title_contains("Problems - LeetCode"))
    # print(f"title is: {browser.title}")

    soup = BeautifulSoup(pageSource, 'lxml')
    if (browser.title == "Problems - LeetCode"):
        
        # page to fetch data
        newSoup = BeautifulSoup(pageSource, 'lxml')
        # fetch the block of questions div
        questionBlock = newSoup.find('div', role='rowgroup')
        # fetch all the questions
        questionList = questionBlock.find_all('div', role='row')

        for question in questionList:
            row = question.find_all('div', role='cell')
            questionName = row[1].find('a').text.split(". ")[1]
            questionUrl = row[1].find('a')['href']
            questionUrl = 'https://leetcode.com' + questionUrl
            questionDifficulty = row[4].find('span').text
            questionNameList.append(questionName)
            questionUrlList.append(questionUrl)
            questionDifficultyList.append(questionDifficulty)
            # print(questionName, questionUrl, questionDifficulty)
        print("********Done*********")
        closeBrowser(browser)

    else:
        print("Page does not exist o connection Failed, status code: ",
              soup.status_code)
    return


def getData():

    try:
        # Opening browser with Headless mode and wait for 2 seconds for page to load
        browser = openBrowser(siteUrl)
        time.sleep(2)
        
        # Fetching the first page data and util the title is "Problems - LeetCode"
        pageSource = browser.page_source
        WebDriverWait(browser, 10).until(EC.title_contains("Problems - LeetCode"))
        soup = BeautifulSoup(pageSource, 'lxml')

        # If title is "Problems - LeetCode" then fetch the data
        if (browser.title == "Problems - LeetCode"):
            
            # Fetching total number of pages
            totalPage = soup.find_all(class_ = "flex items-center justify-center px-3 h-8 rounded select-none focus:outline-none bg-fill-3 dark:bg-dark-fill-3 text-label-2 dark:text-dark-label-2 hover:bg-fill-2 dark:hover:bg-dark-fill-2")
            totalPage = totalPage[-2].text
            totalPage = int(totalPage)
            print(f"Total {totalPage} pages available")
            closeBrowser(browser)

            # Fetching data from each page
            for page in range(1, totalPage + 1):
                print(f"\n********Fetching Page {page}********")
                pageUrl = siteUrl + '?page=' + str(page)
                fetchPageData(pageUrl)

            # All fetched data and now creating excel sheet with the data
            print("*****Done all pages*****")
            print(f"Total {questionNameList.__len__()} questions fetched")
            xcelSheet()

        else:
            print("Connection Failed")
            return

    except Exception as e:
        print("Some error occured, error: ", e)
        return




if __name__ == "__main__":
    getData()