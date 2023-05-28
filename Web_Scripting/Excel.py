from openpyxl import Workbook


def xcelSheet(name , sheet , data):
    wb = Workbook()
    sheet1 = wb.create_sheet(sheet)
    
    for i in range(len(data)):
        for j in range(len(data[i])):
            sheet1.cell(i+1, j+1, data[i][j])

    wb.save(name + '.xlsx')
    wb.close()
    print("\n****Excel sheet created***** ")
    return 
